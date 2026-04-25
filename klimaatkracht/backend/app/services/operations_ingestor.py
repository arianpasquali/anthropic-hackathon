"""Periodic operational data ingestion from food banks.

Accepts CSV or Excel uploads with category-level kg breakdown plus operational
metadata (households served, distribution count, transport km, refrigeration
kWh, operational cost). Validates against business rules and returns clean
records the rest of the pipeline consumes.

The chapter coordinator's experience is the design constraint: errors must
reference specific row numbers and the offending value, so a non-developer
can fix their spreadsheet and retry without engineering help.
"""

import csv
from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO, StringIO, TextIOBase
from typing import BinaryIO, Iterable

CANONICAL_CATEGORIES: tuple[str, ...] = (
    "fresh_produce",
    "bread_bakery",
    "dairy",
    "meat_processed",
    "ready_meals",
    "dry_goods",
    "canned",
    "frozen",
)

KNOWN_CHAPTER_IDS: frozenset[str] = frozenset({
    "VB-AMS-OOST",
    "VB-RDM-ZUID",
    "VB-LEI-CTR",
    "VB-FRL-NRD",
    "VB-EHV-CTR",
})


class ValidationError(Exception):
    """Raised when an uploaded record fails business-rule validation."""


@dataclass
class OperationsRecord:
    chapter_id: str
    period_start: date
    period_end: date
    category_breakdown: dict[str, float]
    households_served: int
    distribution_count: int | None = None
    transport_km: float | None = None
    refrigeration_kwh: float | None = None
    operational_cost_eur: float | None = None
    submission_method: str = "csv_upload"
    raw_input_url: str | None = None

    @property
    def total_kg(self) -> float:
        return sum(self.category_breakdown.values())


def _parse_date(value: str, *, row: int, column: str) -> date:
    try:
        return datetime.strptime(value.strip(), "%Y-%m-%d").date()
    except ValueError as exc:
        raise ValidationError(
            f"Row {row}: column '{column}' has invalid date '{value}', expected YYYY-MM-DD"
        ) from exc


def _parse_number(value: str, *, row: int, column: str, allow_blank: bool = False) -> float | None:
    if value is None or value == "":
        if allow_blank:
            return None
        raise ValidationError(f"Row {row}: column '{column}' is required")
    try:
        return float(value)
    except ValueError as exc:
        raise ValidationError(
            f"Row {row}: column '{column}' has non-numeric value '{value}'"
        ) from exc


def _build_record_from_row(
    row_number: int,
    row: dict[str, str],
    *,
    submission_method: str,
) -> OperationsRecord:
    chapter_id = (row.get("chapter_id") or "").strip()
    if chapter_id not in KNOWN_CHAPTER_IDS:
        raise ValidationError(f"Row {row_number}: unknown chapter '{chapter_id}'")

    period_start = _parse_date(row["period_start"], row=row_number, column="period_start")
    period_end = _parse_date(row["period_end"], row=row_number, column="period_end")
    if period_end < period_start:
        raise ValidationError(
            f"Row {row_number}: period_end {period_end} precedes period_start {period_start}"
        )

    breakdown: dict[str, float] = {}
    for category in CANONICAL_CATEGORIES:
        col = f"{category}_kg"
        if col not in row:
            raise ValidationError(f"Row {row_number}: missing column '{col}'")
        kg = _parse_number(row[col], row=row_number, column=col)
        if kg < 0:
            raise ValidationError(
                f"Row {row_number}: column '{col}' has negative value {kg}"
            )
        breakdown[category] = kg

    households_served_raw = _parse_number(
        row["households_served"], row=row_number, column="households_served"
    )
    if households_served_raw < 0:
        raise ValidationError(
            f"Row {row_number}: 'households_served' has negative value {households_served_raw}"
        )

    return OperationsRecord(
        chapter_id=chapter_id,
        period_start=period_start,
        period_end=period_end,
        category_breakdown=breakdown,
        households_served=int(households_served_raw),
        distribution_count=_optional_int(row.get("distribution_count"), row_number, "distribution_count"),
        transport_km=_optional_float(row.get("transport_km"), row_number, "transport_km"),
        refrigeration_kwh=_optional_float(row.get("refrigeration_kwh"), row_number, "refrigeration_kwh"),
        operational_cost_eur=_optional_float(
            row.get("operational_cost_eur"), row_number, "operational_cost_eur"
        ),
        submission_method=submission_method,
    )


def _optional_int(raw: str | None, row: int, column: str) -> int | None:
    parsed = _parse_number(raw or "", row=row, column=column, allow_blank=True)
    return int(parsed) if parsed is not None else None


def _optional_float(raw: str | None, row: int, column: str) -> float | None:
    return _parse_number(raw or "", row=row, column=column, allow_blank=True)


def parse_operations_csv(stream: TextIOBase) -> list[OperationsRecord]:
    """Parse a CSV upload into validated OperationsRecord instances.

    Each row's row_number in errors is 1-indexed and includes the header,
    so the first data row is row 2 (matches what spreadsheet apps display).
    """
    reader = csv.DictReader(stream)
    records: list[OperationsRecord] = []
    for offset, row in enumerate(reader):
        row_number = offset + 2  # account for header line
        records.append(
            _build_record_from_row(row_number, row, submission_method="csv_upload")
        )
    return records


def parse_operations_xlsx(stream: BinaryIO | BytesIO) -> list[OperationsRecord]:
    """Parse a .xlsx upload. Same validation rules as CSV."""
    from openpyxl import load_workbook

    wb = load_workbook(stream, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [(cell or "") for cell in next(rows_iter)]
    except StopIteration:
        return []

    records: list[OperationsRecord] = []
    for offset, raw_row in enumerate(rows_iter):
        if raw_row is None or all(c is None for c in raw_row):
            continue
        row_number = offset + 2
        row_dict = {
            str(header[i]): _stringify(raw_row[i]) if i < len(raw_row) else ""
            for i in range(len(header))
        }
        records.append(
            _build_record_from_row(row_number, row_dict, submission_method="spreadsheet")
        )
    return records


def _stringify(cell: object) -> str:
    if cell is None:
        return ""
    if isinstance(cell, datetime):
        return cell.date().isoformat()
    if isinstance(cell, date):
        return cell.isoformat()
    return str(cell)


def validate_operations_record(
    record: OperationsRecord,
    *,
    existing: Iterable[OperationsRecord],
) -> None:
    """Check business invariants that span existing records.

    - Date range must not overlap any existing record for the same chapter.
    - Categories must be drawn from the canonical set.
    """
    for cat in record.category_breakdown:
        if cat not in CANONICAL_CATEGORIES:
            raise ValidationError(f"Unknown category '{cat}'")

    for prior in existing:
        if prior.chapter_id != record.chapter_id:
            continue
        if record.period_start <= prior.period_end and record.period_end >= prior.period_start:
            raise ValidationError(
                f"Overlapping period for chapter {record.chapter_id}: "
                f"existing {prior.period_start}–{prior.period_end} vs "
                f"new {record.period_start}–{record.period_end}"
            )
